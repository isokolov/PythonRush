import pandas as pd
from openpyxl import load_workbook
import shutil

# Указываем пути к CSV и Excel файлам
csv_file_path = 'employees.csv'  # Путь к CSV файлу
excel_file_path = 'employees_data.xlsx'  # Путь к существующему Excel файлу

# Чтение данных из CSV файла
csv_data = pd.read_csv(csv_file_path)

# Создание резервной копии Excel файла
backup_path = excel_file_path.replace('.xlsx', '_backup.xlsx')
shutil.copy(excel_file_path, backup_path)

# Загрузка существующего Excel файла
workbook = load_workbook(excel_file_path)
sheet = workbook.active

# Определение начальной строки для добавления новых данных
start_row = sheet.max_row + 1

# Добавление данных из CSV в Excel
for index, row in csv_data.iterrows():
    for col_num, value in enumerate(row, start=1):
        # Записываем значение в соответствующую ячейку
        sheet.cell(row=start_row + index, column=col_num, value=value)

# Сохраняем изменения в Excel файле
workbook.save(excel_file_path)
